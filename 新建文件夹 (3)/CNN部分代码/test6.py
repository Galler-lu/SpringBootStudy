import numpy as np
import matplotlib.pyplot as plt
import torch
import xlwt
import openpyxl
from torch.utils.data import Dataset, DataLoader
import torch.optim as optim

my_list = []
wb = openpyxl.load_workbook('test1.xlsx')
ws = wb['Sheet1']
maxrows = ws.max_row  # 获取最大行
for i in range(maxrows - 1):
    temp_list = []
    for each in ws.iter_cols(min_row=2):
        temp_list.append(each[i].value)
    my_list.append(temp_list)

from sklearn.preprocessing import MinMaxScaler


# dataSet加载数据集
class DiabetesDataset(Dataset):
    def __init__(self, my_list):
        xy = np.array(my_list)
        self.len1 = xy.shape[0]
        temp_list = []
        self.y_data = torch.from_numpy(xy[:, [-1]])

        self.transfer = MinMaxScaler()
        xy = self.transfer.fit_transform(xy[:, :-1])

        for i in range(len(xy)):
            temp_list.append([list(xy[i, :6]), list(xy[i, 6:12]), list(xy[i, 12:18])])
        self.x_data = torch.from_numpy(np.array(temp_list))

    def __getitem__(self, index):
        return self.x_data[index], self.y_data[index]

    def __len__(self):
        return self.len1


dataset = DiabetesDataset(my_list)
train_loader = DataLoader(dataset=dataset, batch_size=2, shuffle=True, num_workers=0)


# 卷积神经网络模型
class NetModel1(torch.nn.Module):
    def __init__(self):
        super(NetModel1, self).__init__()
        self.conv_ghwa_1 = torch.nn.Conv1d(3, 16, kernel_size=3, padding=1)
        self.conv_ghwa_2 = torch.nn.Conv1d(16, 32, kernel_size=3, padding=1)
        self.relu = torch.nn.ReLU()
        self.averagePooling1 = torch.nn.AvgPool1d(6)
        self.linear = torch.nn.Linear(32, 1)
        self.softmax = torch.nn.Softmax()
        self.sigmoid = torch.nn.Sigmoid()

    def forward(self, x):
        x = self.relu(self.conv_ghwa_1(x))
        x = self.relu(self.conv_ghwa_2(x))
        x = self.averagePooling1(x)
        x = x.view(-1, 32)
        return self.sigmoid(self.linear(x))


model = NetModel1()

criterion = torch.nn.BCELoss()
optimizer = torch.optim.SGD(model.parameters(), lr=0.01, momentum=0.5)

for epoch in range(100):
    # Forward
    for batch_idx, data in enumerate(train_loader, 0):
        x_data, y_data = data
        y_pred = model(x_data.float())
        # if epoch % 100 == 99:
        #     print('y_pred为：', y_pred, y_pred.size())
        y_pred = y_pred.to(torch.float32)
        y_data = y_data.to(torch.float32)
        loss = criterion(y_pred, y_data)
        if epoch % 100 == 99:
            print(epoch, loss.item())
        optimizer.zero_grad()
        loss.backward()
        optimizer.step()
