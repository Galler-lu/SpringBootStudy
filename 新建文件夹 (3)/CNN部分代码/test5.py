import numpy as np
import torch
import xlwt
import openpyxl

my_list = []
wb = openpyxl.load_workbook('test1.xlsx')
ws = wb['Sheet1']
maxrows = ws.max_row  # 获取最大行
for i in range(maxrows - 1):
    temp_list = []
    for each in ws.iter_cols(min_row=2):
        temp_list.append(each[i].value)
    my_list.append(temp_list)

xy = np.array(my_list)
len1 = xy.shape[0]
temp_list = []

from sklearn.preprocessing import MinMaxScaler

y_data = torch.from_numpy(xy[:, [-1]])

transfer = MinMaxScaler()
xy = transfer.fit_transform(xy[:, :-1])

for i in range(len(xy)):
    temp_list.append([list(xy[i, :6]), list(xy[i, 6:12]), list(xy[i, 12:18])])

x_data = torch.from_numpy(np.array(temp_list))


# 卷积神经网络模型
class NetModel1(torch.nn.Module):
    def __init__(self):
        super(NetModel1, self).__init__()
        self.conv_ghwa_1 = torch.nn.Conv1d(3, 16, kernel_size=3, padding=1)
        self.conv_ghwa_2 = torch.nn.Conv1d(16, 32, kernel_size=3, padding=1)
        self.relu = torch.nn.ReLU()
        self.averagePooling1 = torch.nn.AvgPool1d(6)
        self.linear = torch.nn.Linear(32, 1)
        self.softmax = torch.nn.Softmax(dim=1)
        self.sigmoid = torch.nn.Sigmoid()

    def forward(self, x):
        x = self.relu(self.conv_ghwa_1(x))
        x = self.relu(self.conv_ghwa_2(x))
        x = self.averagePooling1(x)
        x = x.view(-1, 32)
        return self.sigmoid(self.linear(x))


model = NetModel1()

criterion = torch.nn.BCELoss()
optimizer = torch.optim.SGD(model.parameters(), lr=0.01, momentum=0.5)


def train(number):
    for epoch in range(number):
        y_pred = model(x_data.float())
        y_pred = y_pred.to(torch.float32)
        y_data1 = y_data.to(torch.float32)

        loss = criterion(y_pred, y_data1)

        # print(epoch, loss.item())
        optimizer.zero_grad()
        loss.backward()
        optimizer.step()





def test():
    correct = 0
    total = 0
    outputs = model(x_data.float())
    outputs = outputs.to(torch.float32)
    y_data2 = y_data.to(torch.float32)
    output_prediction = (outputs > 0.5).float()
    print(output_prediction)
    total += y_data2.size(0)
    correct += torch.sum(output_prediction == y_data2.data)
    print('total为：', total)
    print('correct为', correct)
    print('Accuracy on test set: %d %%' % (100 * correct / total))


train(1000)
test()
