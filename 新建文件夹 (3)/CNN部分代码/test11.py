import random

import numpy as np
import torch
import xlwt
import openpyxl
import matplotlib.pyplot as plt

import os
from sklearn.preprocessing import MinMaxScaler
from sklearn.preprocessing import StandardScaler

# 正常数据
my_list_normal = []
wb = openpyxl.load_workbook('People7Normal.xlsx')
ws = wb['Sheet1']
maxrows = ws.max_row  # 获取最大行
for i in range(maxrows - 1):
    temp_list = []
    for each in ws.iter_cols(min_row=1):
        if each[i].value == 0 or each[i].value:
            temp_list.append(each[i].value)
    my_list_normal.append(temp_list)
random.shuffle(my_list_normal)
# xy_normal = np.array(my_list_normal)
# len1 = xy.shape[0]
temp_list = []

# 发病数据
my_list_sick = []
wb = openpyxl.load_workbook('People7Sick.xlsx')
ws = wb['Sheet1']
maxrows = ws.max_row  # 获取最大行
for i in range(maxrows - 1):
    temp_list = []
    for each in ws.iter_cols(min_row=1):
        if each[i].value == 0 or each[i].value:
            temp_list.append(each[i].value)
    my_list_sick.append(temp_list)
random.shuffle(my_list_sick)

# xy_sick = np.array(my_list_sick)
# len1 = xy.shape[0]
temp_list = []
# 划分训练集和测试集7:3分
train_sick = my_list_sick[:int(len(my_list_sick) * 0.7)]
train_normal = my_list_normal[:len(train_sick)]
test_sick = my_list_sick[len(train_sick):]
test_normal = my_list_normal[len(train_normal):len(train_normal) + len(test_sick)]
train_normal.extend(train_sick)
test_normal.extend(test_sick)

train_data = train_normal
test_data = test_normal

# 打乱训练集和测试集
random.shuffle(train_data)
random.shuffle(test_data)

xy = np.array(train_data)

y_data = torch.from_numpy(xy[:, [-1]])

# transfer = MinMaxScaler()
transfer = StandardScaler()
xy = transfer.fit_transform(xy[:, :-1])

for i in range(len(xy)):
    temp_list.append([list(xy[i, :6]), list(xy[i, 6:12]), list(xy[i, 12:18])])

x_data = torch.from_numpy(np.array(temp_list))


#GoogLeNet模型

