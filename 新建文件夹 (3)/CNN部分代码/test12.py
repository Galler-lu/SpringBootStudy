import random
import time
import pandas as pd
import numpy as np
import torch
import xlwt
import openpyxl
import matplotlib.pyplot as plt
import numpy as np
import os
from sklearn.preprocessing import MinMaxScaler
from sklearn.preprocessing import StandardScaler

print('开始时间：', time.strftime('%Y-%m-%d %X', time.localtime(time.time())))

# 读取正常数据
normal = pd.read_excel('People_7_Normal.xlsx', header=None)
data_ls_normal = [normal.columns.tolist()] + normal.values.tolist()
my_list_normal = np.array(data_ls_normal)
my_list_normal = my_list_normal[1:]

# 读取发病数据
sick = pd.read_excel('People_7_Sick.xlsx', header=None)
data_ls_sick = [sick.columns.tolist()] + sick.values.tolist()
my_list_sick = np.array(data_ls_sick)
my_list_sick = my_list_sick[1:]
random.shuffle(my_list_normal)
random.shuffle(my_list_sick)

# # 正常数据
# my_list_normal = []
# wb = openpyxl.load_workbook('People7Normal.xlsx')
# ws = wb['Sheet1']
# maxrows = ws.max_row  # 获取最大行
# for i in range(maxrows - 1):
#     temp_list = []
#     for each in ws.iter_cols(min_row=1):
#         if each[i].value == 0 or each[i].value:
#             temp_list.append(each[i].value)
#     my_list_normal.append(temp_list)
# random.shuffle(my_list_normal)
# # xy_normal = np.array(my_list_normal)
# # len1 = xy.shape[0]
# temp_list = []
# print('开始时间0：', time.strftime('%Y-%m-%d %X', time.localtime(time.time())))
# # 发病数据
# my_list_sick = []
# wb = openpyxl.load_workbook('People7Sick.xlsx')
# ws = wb['Sheet1']
# maxrows = ws.max_row  # 获取最大行
# for i in range(maxrows - 1):
#     temp_list = []
#     for each in ws.iter_cols(min_row=1):
#         if each[i].value == 0 or each[i].value:
#             temp_list.append(each[i].value)
#     my_list_sick.append(temp_list)
# random.shuffle(my_list_sick)

print('开始时间1：', time.strftime('%Y-%m-%d %X', time.localtime(time.time())))

xy_sick = my_list_sick
temp_list = []
# 划分训练集和测试集7:3分
train_sick = list(my_list_sick[:int(len(my_list_sick) * 0.7)])
train_normal = list(my_list_normal[:len(train_sick)])
test_sick = list(my_list_sick[len(train_sick):])
test_normal = list(my_list_normal[len(train_normal):len(train_normal) + len(test_sick)])
train_normal.extend(train_sick)
test_normal.extend(test_sick)

train_data = train_normal
test_data = test_normal

# 打乱训练集和测试集
random.shuffle(train_data)
random.shuffle(test_data)

xy = np.array(train_data)


y_data = torch.from_numpy(xy[:, [-1]])

# transfer = MinMaxScaler()
transfer = StandardScaler()
xy = transfer.fit_transform(xy[:, :-1])

for i in range(len(xy)):
    temp_list.append([list(xy[i, :6]), list(xy[i, 6:12]), list(xy[i, 12:18])])

x_data = torch.from_numpy(np.array(temp_list))


# #
# #
# # 卷积神经网络模型
# class NetModel1(torch.nn.Module):
#     def __init__(self):
#         super(NetModel1, self).__init__()
#         self.conv_ghwa_1 = torch.nn.Conv1d(3, 16, kernel_size=3, padding=1)
#         self.conv_ghwa_2 = torch.nn.Conv1d(16, 32, kernel_size=3, padding=1)
#         self.relu = torch.nn.ReLU()
#         self.averagePooling1 = torch.nn.AvgPool1d(6)
#         self.linear = torch.nn.Linear(32, 1)
#         self.softmax = torch.nn.Softmax(dim=1)
#         self.sigmoid = torch.nn.Sigmoid()
#
#     def forward(self, x):
#         x = self.relu(self.conv_ghwa_1(x))
#         x = self.relu(self.conv_ghwa_2(x))
#         x = self.averagePooling1(x)
#         x = x.view(-1, 32)
#         return self.sigmoid(self.linear(x))
#
#
# model = NetModel1()
#
# criterion = torch.nn.BCELoss()
# optimizer = torch.optim.SGD(model.parameters(), lr=0.3, momentum=0.5)
#
# loss1 = []
#
# xy1 = np.array(test_data)
# y_data1 = torch.from_numpy(xy1[:, [-1]])
#
# # transfer1 = Normalizer()
# # transfer1 = MinMaxScaler()
# xy1 = transfer.fit_transform(xy1[:, :-1])
# temp_list1 = []
# for i in range(len(xy1)):
#     temp_list1.append([list(xy1[i, :6]), list(xy1[i, 6:12]), list(xy1[i, 12:18])])
#
# x_data1 = torch.from_numpy(np.array(temp_list1))
#
# accuracy = []
#
#
# def test():
#     correct = 0
#     total = 0
#     correct1 = 0
#     correct0 = 0
#     outputs = model(x_data1.float())
#     outputs = outputs.to(torch.float32)
#     y_data2 = y_data1.to(torch.float32)
#     output_prediction = (outputs > 0.5).float()
#     # print(output_prediction)
#     # list = tensor.numpy().tolist()
#     total += y_data2.size(0)
#     correct += torch.sum(output_prediction == y_data2.data)
#     correct1 += torch.sum(output_prediction == 1)
#     correct0 += torch.sum(output_prediction == 0)
#     print('total为：', total)
#     print('correct为', correct)
#     print('correct1为：', correct1)
#     print('correct0为：', correct0)
#     print('Accuracy on test set: %d %%' % (100 * correct / total))
#     accuracy.append(correct / total)
#
#
# def train(number):
#     # lossSum = 0
#     lossValue = 0
#     for epoch in range(number):
#         y_pred = model(x_data.float())
#         y_pred = y_pred.to(torch.float32)
#         y_data1 = y_data.to(torch.float32)
#
#         loss = criterion(y_pred, y_data1)
#
#         lossValue = loss.item()
#         print(epoch, lossValue)
#
#         # lossSum = lossSum+loss.item()
#
#         optimizer.zero_grad()
#         loss.backward()
#         optimizer.step()
#         if epoch % 100 == 99:
#             print("第1个*************************************这是第", epoch, "在测试集上的结果")
#             test()
#         loss1.append(lossValue)
#
#
# # 保存方式一
#
# # print(model.state_dict())
# # torch.save(model.module.state_dict(), 'model1.pth')
# # 保存方式二
# # torch.save(model, 'model2.pth')
# # 读取方式一
# # model = torch.load('model2.pth')
# # 读取方式二
# # model = NetModel1()
# # model.load_state_dict(torch.load('model1.pth'))
#
# number = 1000
# train(number)
# min_loss = min(loss1)
# min_loss_index = loss1.index(min_loss)
# model_name = 'model迭代次数为_' + str(number) + '最小损失为_' + str(min_loss) + '对应索引为_' + str(min_loss_index) + '.pth'
# torch.save(model.state_dict(), model_name)
# print('开始时间2：', time.strftime('%Y-%m-%d %X', time.localtime(time.time())))
# # print('开始时间3：', time.strftime('%Y-%m-%d %X', time.localtime(time.time())))
# plt.subplot(1, 2, 1)
# plt.plot(loss1, label='loss')
# plt.subplot(1, 2, 2)
# plt.plot(accuracy, label='testAccuracy')
# plt.legend()
# plt.show()
