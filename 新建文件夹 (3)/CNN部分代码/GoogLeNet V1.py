import random

import numpy as np
import torch
import xlwt
import openpyxl
import matplotlib.pyplot as plt
import torch.nn.functional as F
import os
from sklearn.preprocessing import MinMaxScaler
from sklearn.preprocessing import StandardScaler

# 正常数据
my_list_normal = []
wb = openpyxl.load_workbook('NormalSet.xlsx')
ws = wb['Sheet1']
maxrows = ws.max_row  # 获取最大行
for i in range(maxrows - 1):
    temp_list = []
    for each in ws.iter_cols(min_row=1):
        if each[i].value == 0 or each[i].value:
            temp_list.append(each[i].value)
    my_list_normal.append(temp_list)
random.shuffle(my_list_normal)
# xy_normal = np.array(my_list_normal)
# len1 = xy.shape[0]
temp_list = []

# 发病数据
my_list_sick = []
wb = openpyxl.load_workbook('SickSet.xlsx')
ws = wb['Sheet1']
maxrows = ws.max_row  # 获取最大行
for i in range(maxrows - 1):
    temp_list = []
    for each in ws.iter_cols(min_row=1):
        if each[i].value == 0 or each[i].value:
            temp_list.append(each[i].value)
    my_list_sick.append(temp_list)
random.shuffle(my_list_sick)

# xy_sick = np.array(my_list_sick)
# len1 = xy.shape[0]
temp_list = []
# 划分训练集和测试集7:3分
train_sick = my_list_sick[:int(len(my_list_sick) * 0.7)]
train_normal = my_list_normal[:len(train_sick)]
test_sick = my_list_sick[len(train_sick):]
test_normal = my_list_normal[len(train_normal):len(train_normal) + len(test_sick)]
train_normal.extend(train_sick)
test_normal.extend(test_sick)

train_data = train_normal
test_data = test_normal

# 打乱训练集和测试集
random.shuffle(train_data)
random.shuffle(test_data)

xy = np.array(train_data)

y_data = torch.from_numpy(xy[:, [-1]])

# transfer = MinMaxScaler()
transfer = StandardScaler()
xy = transfer.fit_transform(xy[:, :-1])

for i in range(len(xy)):
    temp_list.append([list(xy[i, :6]), list(xy[i, 6:12]), list(xy[i, 12:18])])

x_data = torch.from_numpy(np.array(temp_list))


# Inception模型
class InceptionA(torch.nn.Module):
    def __init__(self, in_channels):
        self.branch1x1 = torch.nn.Conv1d(in_channels, 16, kernel_size=1, padding=1)

        self.branch5x5_1 = torch.nn.Conv1d(in_channels, 16, kernel_size=1, padding=0)
        self.branch5x5_2 = torch.nn.Conv1d(16, 24, kernel_size=5, padding=2)

        self.branch3x3_1 = torch.nn.Conv1d(in_channels, 16, kernel_size=1, padding=0)
        self.branch3x3_2 = torch.nn.Conv1d(16, 24, kernel_size=3, padding=1)
        self.branch3x3_3 = torch.nn.Conv1d(24, 24, kernel_size=3, padding=1)

        self.branch_pool = torch.nn.Conv1d(in_channels, 24, kernel_size=1)

    def forward(self, x):
        branch1x1 = self.branch1x1(x)

        branch5x5 = self.branch5x5_1(x)
        branch5x5 = self.branch5x5_2(branch5x5)

        branch3x3 = self.branch3x3_1(x)
        branch3x3 = self.branch3x3_2(branch3x3)
        branch3x3 = self.branch3x3_3(branch3x3)

        branch_pool = F.avg_pool1d(x, kernel_size=3, stride=1, padding=1)
        branch_pool = self.branch_pool(branch_pool)

        outputs = [branch1x1, branch5x5, branch3x3, branch_pool]
        return torch.cat(outputs, dim=1)


# GoogLeNet V1模型
class Net(torch.nn.Module):
    def __init__(self):
        super(Net, self).__init__()
        self.conv1 = torch.nn.Conv1d(1, 10, kernel_size=5)
        self.conv2 = torch.nn.Conv1d(88, 20, kernel_size=5)

        self.incep1 = InceptionA(in_channels=10)
        self.incep2 = InceptionA(in_channels=20)

        self.mp = torch.nn.MaxPool1d(2)
        self.fc = torch.nn.Linear(1408, 10)

    def forward(self, x):
        in_size = x.size(0)
        x = F.relu(self.mp(self.conv1(x)))
        x = self.incep1(x)
        x = F.relu(self.mp(self.conv2(x)))
        x = self.incep2(x)
        x = x.view(in_size, -1)
        x = self.fc(x)
        return x
