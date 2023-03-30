import random
import numpy as np
import torch
import xlwt
import openpyxl
import os
from sklearn.preprocessing import MinMaxScaler

if __name__ == '__main__':
    # 卷积神经网络模型
    class NetModel1(torch.nn.Module):
        def __init__(self):
            super(NetModel1, self).__init__()
            self.conv_ghwa_1 = torch.nn.Conv1d(3, 16, kernel_size=3, padding=1)
            self.conv_ghwa_2 = torch.nn.Conv1d(16, 32, kernel_size=3, padding=1)
            self.relu = torch.nn.ReLU()
            self.averagePooling1 = torch.nn.AvgPool1d(6)
            self.linear = torch.nn.Linear(32, 1)
            self.softmax = torch.nn.Softmax(dim=1)
            self.sigmoid = torch.nn.Sigmoid()

        def forward(self, x):
            x = self.relu(self.conv_ghwa_1(x))
            x = self.relu(self.conv_ghwa_2(x))
            x = self.averagePooling1(x)
            x = x.view(-1, 32)
            return self.sigmoid(self.linear(x))


    # 读取方式二
    model = NetModel1()
    model.load_state_dict(torch.load('model1.pth'))

    my_list1 = []
    wb = openpyxl.load_workbook('test_2.xlsx')
    ws = wb['Sheet1']
    maxrows = ws.max_row  # 获取最大行
    for i in range(maxrows - 1):
        temp_list1 = []
        for each in ws.iter_cols(min_row=2):
            if each[i].value == 0 or each[i].value:
                temp_list1.append(each[i].value)
        my_list1.append(temp_list1)

    xy1 = np.array(my_list1)
    len11 = xy1.shape[0]
    temp_list1 = []

    from sklearn.preprocessing import MinMaxScaler

    print(my_list1)
    y_data = torch.from_numpy(xy1[:, [-1]])

    transfer1 = MinMaxScaler()
    xy1 = transfer1.fit_transform(xy1[:, :-1])

    for i in range(len(xy1)):
        temp_list1.append([list(xy1[i, :6]), list(xy1[i, 6:12]), list(xy1[i, 12:18])])

    x_data = torch.from_numpy(np.array(temp_list1))


    def test():
        correct = 0
        total = 0
        correct1 = 0
        correct0 = 0
        outputs = model(x_data.float())
        outputs = outputs.to(torch.float32)
        y_data2 = y_data.to(torch.float32)
        output_prediction = (outputs > 0.5).float()

        print(output_prediction)

        out = output_prediction.numpy().tolist()
        import tkinter as tk
        a11 = list(tk._flatten(out))
        print(a11)
        y_pred_bbb = [int(a) for a in a11]

        # for four_ck in range(30):
        #     co = 5 + four_ck  # 连续多少个1判断发病为1
        #     m = 0  # 连续出现1的个数
        #     k = 0  # 记录上一次判断为发病的时间10分钟以内
        #     xgl = []  # 记录报警事件
        for i in range(len(y_pred_bbb)):
            res = y_pred_bbb[i]
            # if i - k < 600 and k != 0:
            #     continue

            # if res == 0:
            #     m = 0
            #     continue
            if res == 1:
                # m += 1
                # if m == co:
                #     k = i
                #     m = 0
                print("报警了", i)

        total += y_data2.size(0)
        correct += torch.sum(output_prediction == y_data2.data)
        correct1 += torch.sum(output_prediction == 1)
        correct0 += torch.sum(output_prediction == 0)
        print('total为：', total)
        print('correct为', correct)
        print('correct1为：', correct1)
        print('correct0为：', correct0)
        print('Accuracy on test set: %d %%' % (100 * correct / total))


    test()
